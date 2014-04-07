#!/usr/bin/python
#-*- coding: utf-8 -*-


import os
import shutil
from struct import *  
import sys
from openpyxl.reader.excel import load_workbook



def convert_sheet_to_code(ws,code):
#保存一个代码结构块
    cols = ws.get_highest_row()
    if cols > 2:
        cols = 2
    for rx in range(0,cols):  
        temp_list = []       
        for ry in range(2,ws.get_highest_column()):           
            temp_list.append(ws.cell(row = rx,column = ry).value)       
        code[str(rx)] = temp_list  


def convert_sheet_to_dict(ws,dict):
    #把数据存到字典中 
    for rx in range(3,ws.get_highest_row()):  
        temp_list = []  
        key = ws.cell(row = rx,column = 0).value 
        #print key
        for ry in range(2,ws.get_highest_column()):
            #if  ry == 1: #desc不导出
            #   continue
            temp_list.append(unicode(ws.cell(row = rx,column = ry).value))
        #print temp_list
        dict[str(key)] = temp_list

    

def write_Plist_code(path_code,codes):
	#保持代码结构
	newPlistCode = open(path_code,"wb") 
	newPlistCode.write('''<?xml version="1.0" encoding="UTF-8"?>\n''') 
	newPlistCode.write('''<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "./PropertyList-1.0.dtd">\n''')
	newPlistCode.write('''<plist version="1.0">\n''')
	newPlistCode.write("<dict>\n")

	count = 0 
	for key in codes.keys():
		code_list = codes[key]
		if count == 0:
			newPlistCode.write("<name>\n")
		else:
			newPlistCode.write("<type>\n")
		for data in code_list:
			newPlistCode.write("<string>")
			newPlistCode.write(str(data))
			newPlistCode.write("</string>\n")
		if count == 0:
			newPlistCode.write("</name>\n")
		else:
			newPlistCode.write("</type>\n")
		count += 1
	newPlistCode.write("</dict>\n")
	newPlistCode.write("</plist>\n")
	newPlistCode.close()
	
def write_Plist_data(path_plist,dict):
	#把字典转换为plist
    print(path_plist)
    newPlistFile = open(path_plist,"wb") 
    #newPlistFile.write('''<?xml version="1.0" encoding="UTF-8"?>\n''') 
    #newPlistFile.write('''<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "./PropertyList-1.0.dtd">\n''')
    #newPlistFile.write('''<plist version="1.0">\n''')
    #newPlistFile.write("<dict>\n")    
    content = u''
    for key in dict.keys():
        content += (str(key))        
        temp_list = dict[key]
        for data in temp_list:   
        	content+=u"\t"
        	content+=unicode(data)
        content+=u"\n" 
    newPlistFile.write(content.encode("utf-8"))
    newPlistFile.close()



def convert_excel_to_plist(path,path_plist,path_code):
	wb = load_workbook(filename = path)
    # 建立存储数据的字典，用name作为字典key   
	dict = {} 
	codes = {}	
    #取第一张表  
	sheetnames = wb.get_sheet_names()
	name = sheetnames[0]
	ws = wb.get_sheet_by_name(name)
	convert_sheet_to_dict(ws,dict)
	convert_sheet_to_code(ws,codes)
	'''
	for name in sheetnames:
		ws = wb.get_sheet_by_name(name)
		if name == "function":
			convert_sheet_to_dict(ws,function_dict)
		elif name == "price":
			convert_sheet_to_dict(ws,price_dict)
		elif name == "cooldown":
			convert_sheet_to_dict(ws,cooldown_dict)
	'''
    
	write_Plist_data(path_plist,dict)
	write_Plist_code(path_code,codes)
    
	
	
	    


dir = os.getcwd()
path = os.path.join(dir,"Public")

#设置当前目录路径并开始遍历
for root,dirs,files in os.walk(path,topdown=True):
    for name in files:
        if root.find('.svn') < 0:
            if name.find('.xlsx') > 0:
                name_excel = name
                temp_str = name_excel.split('.')
                name_plist = temp_str[0] + '.txt'
                name_code = temp_str[0] + '.code'
                print name_plist
                path_plist = os.path.join(path,name_plist)
                path_code = os.path.join(path,name_code)
                #print path_plist
                path_excel = os.path.join(path,name_excel)
                #读取Excel并转换为XML
                convert_excel_to_plist(path_excel,path_plist,path_code)
