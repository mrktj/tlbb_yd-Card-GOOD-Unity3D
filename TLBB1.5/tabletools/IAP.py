#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import shutil
from struct import *  
import sys
from openpyxl.reader.excel import load_workbook


def convert_sheet_to_dict(ws,files):
    #把数据存到字典中
    newPlistCode = open(files,"wb")
    newPlistCode.write('''<product bid="sevenseaocean" test="false">\n''')
    
    for rx in range(3,ws.get_highest_row()):        
        newPlistCode.write('''<purchase id="''')
        newPlistCode.write(ws.cell(row = rx,column = 0).value)
        newPlistCode.write('''" name="@:''')
        newPlistCode.write(str(ws.cell(row = rx,column = 2).value))
        newPlistCode.write('''" title="@:''')
        newPlistCode.write(str(ws.cell(row = rx,column = 3).value))
        newPlistCode.write('''" price="''')
        newPlistCode.write(str(ws.cell(row = rx,column = 1).value))
        newPlistCode.write('''" type="''')
        newPlistCode.write(str(ws.cell(row = rx,column = 4).value))
        newPlistCode.write('''" extraImoneyRatio="''')
        newPlistCode.write(str(ws.cell(row = rx,column = 7).value))
        newPlistCode.write('''" imoney="''')
        newPlistCode.write(str(ws.cell(row = rx,column = 5).value))
        newPlistCode.write('''" icon="''')
        newPlistCode.write(ws.cell(row = rx,column = 6).value)
        newPlistCode.write('''" noADTime="''')
        newPlistCode.write(str(ws.cell(row = rx,column = 8).value))
        newPlistCode.write('''"/>\n''')
    newPlistCode.write('''</product>\n''')
    
def convert_excel_to_plist(path,path_plist):
	wb = load_workbook(filename = path)
    # 建立存储数据的字典，用name作为字典key   
	datas = {} 		
    #取第一张表  
	sheetnames = wb.get_sheet_names()
	name = sheetnames[0]
	ws = wb.get_sheet_by_name(name)
	convert_sheet_to_dict(ws,path_plist)	
	
    

dir = os.getcwd()
path = os.path.join(dir,"Datas")
files = os.path.join(path,"AppStoreShop.xlsx");
#读取Excel并转换为XML
convert_excel_to_plist(files,"./so_app.txt")

